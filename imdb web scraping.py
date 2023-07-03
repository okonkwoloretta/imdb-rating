#!/usr/bin/env python
# coding: utf-8

# In[45]:


# Import libraries
from bs4 import BeautifulSoup
import requests
import openpyxl
import os

# Creating an Excel workbook and adding a sheet
excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movies Name', 'Year of Release', 'IMDB Rating'])

# Want to know my current working directory
print("Current Working Directory:", os.getcwd())


# In[47]:


# URL of the IMDb top-rated movies chart
url = 'https://www.imdb.com/chart/top/'

# Adding a User-Agent header to mimic a regular browser request
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

# Fetch the webpage content with the User-Agent header
source = requests.get(url, headers=headers)

try:
    # Check if the request was successful
    source.raise_for_status()

    # Parse the content using BeautifulSoup
    soup = BeautifulSoup(source.text, 'html.parser')
    
    # Extracting movie information
    movies = soup.find('tbody', class_='lister-list').find_all('tr')
  
    # Loop through each movie and print its title and rating
    for movie in movies:
        rank = movie.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]
        name = movie.find('td', class_='titleColumn').find('a').text
        year = movie.find('td', class_='titleColumn').span.text.strip('[]')
        rating = movie.find('td', class_='ratingColumn').strong.text

        print(rank,name,year,rating)
        
        sheet.append([rank,name,year,rating])
        
        # Convert the 'Year of Release' to numeric data type
    try:
        year = int(year)
    except ValueError:
        year = None
    
    # Convert the 'IMDB Rating' to numeric data type
    try:
        rating = float(rating)
    except ValueError:
        rating = None

    # Append cleaned and converted data to the Excel sheet
    sheet.append([rank, name, year, rating])

except requests.exceptions.HTTPError as http_err:
    print(f"HTTP error occurred: {http_err}")
except Exception as err:
    print(f"An error occurred: {err}")
    
    # Save to Excel 
excel.save('IMDB Top Rated Movies.xlsx')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




